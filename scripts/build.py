"""Build the public VetPrices dataset.

Rule: PUBLIC DATA IS BUILT ONLY FROM OFFICIAL LIST PRICES (supplier price lists / public retail
sites). Never from vetmarket.db / pricer.db / clinic Postgres — those carry the clinic's actual
purchase prices, discounts, and sale prices. Fields like customerPriceWithVat are dropped on purpose.

Inputs : _canonical/*.json (from scripts/ts_to_json.js + server pulls), medimarket prices.db
Outputs: data/<type>/<slug>.json, data/index.json, site/downloads/<type>.xlsx, sources/ copies
"""
import json, re, shutil, sqlite3, sys, datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
CAN = ROOT / "_canonical"
DATA = ROOT / "data"
SRC = ROOT / "sources"
DL = ROOT / "Downloads_stub"  # unused
VAT = 1.18
TODAY = datetime.date.today().isoformat()
HOME = Path.home()

# ---------------------------------------------------------------- taxonomy (ported from vetmarket-cli web.py)
TOPICS = [
    ("ears", "אוזניים"), ("eyes", "עיניים"), ("dental", "שיניים ופה"), ("skin", "עור ופרווה"),
    ("joints", "מפרקים וניידות"), ("kidney", "כליה ודרכי שתן"), ("gi", "מערכת עיכול"),
    ("cardio", "לב וכלי דם"), ("endocrine", "הורמונלי / סוכרת"), ("neuro", "נוירולוגיה / התנהגות"),
    ("liver", "כבד"), ("onco", "אונקולוגיה"), ("antibiotic", "אנטיביוטיקה"), ("pain", "כאב ודלקת"),
    ("parasites", "הדברה / טפילים"), ("deworm", "תילוע"), ("vaccine", "חיסונים"),
    ("nutrition", "תזונה ומזון"), ("respiratory", "נשימה"), ("anesthesia", "הרדמה והרגעה"),
    ("equipment", "ציוד ומתכלים"), ("other", "כללי / לא מסווג"),
]
LAB_TOPICS = [
    ("hematology", "המטולוגיה וספירת דם"), ("chemistry", "ביוכימיה ופאנלים"), ("endocrine", "הורמונים"),
    ("serology", "סרולוגיה וזיהומים"), ("pcr", "PCR וגנטיקה"), ("culture", "תרביות"),
    ("pathology", "פתולוגיה וציטולוגיה"), ("urine", "שתן וצואה"), ("coag", "קרישה"),
    ("drugs", "רמות תרופות"), ("other", "אחר"),
]
FOOD_IND_TOPIC = {
    "מחלות כליה": "kidney", "דרכי שתן": "kidney", "מערכת עיכול": "gi", "אלרגיה למזון": "skin",
    "השמנה / ירידה במשקל": "nutrition", "סוכרת": "endocrine", "עור ופרווה": "skin",
    "מפרקים / ניידות": "joints", "התאוששות": "nutrition", "מחלות כבד": "liver", "דלקות פרקים": "joints",
    "אפילפסיה / מערכת עצבים": "neuro", "תחליף חלב": "nutrition", "גורים": "nutrition", "מזון רגיל": "nutrition",
    "לב": "cardio", "בלוטת התריס": "endocrine", "סרטן": "onco", "שיניים": "dental",
}
BE_CANON = {"אזניים": "אוזניים", "אנטיביוטקיה": "אנטיביוטיקה", "קרדיוואסקולר": "קרדיו", "חוטים (המשך)": "חוטים",
            "דרמטולוגיה": "דרמטולוגי", "נוזלים ותמיסות": "נוזלים"}
BE_TOPIC = {
    "אוזניים": "ears", "עיניים": "eyes", "עיניים ושיניים": "eyes", "דנטלי": "dental", "דרמטולוגי": "skin",
    "הדברה": "parasites", "תילוע": "deworm", "חיסונים": "vaccine", "תרכיבים": "vaccine",
    "אנטיביוטיקה": "antibiotic", "NSAID": "pain", "נוגדי דלקת": "pain", "הורמונלי": "endocrine",
    "נוירולוגיה": "neuro", "התנהגות": "neuro", "כבד": "liver", "כליות ושתן": "kidney",
    "מערכת עיכול": "gi", "מערכת עיכול - ארנבונים": "gi", "מפרקים": "joints", "קרדיו": "cardio",
    "ויטמינים": "nutrition", "מינרלים ושומן": "nutrition", "תחליפי חלב": "nutrition", "חטיפים למתן כדורים": "nutrition",
    "הרדמה": "anesthesia",
}
VM_TOPIC = {
    "מזון רפואי": "nutrition", "בדיקות מעבדה": "equipment", "אנטיביוטיקה": "antibiotic", "חומרים מתכלים": "equipment",
    "עיניים ואוזניים": None, "הדברה": "parasites", "מערכת עיכול": "gi", "דרמטולוגיה / אלרגיה": "skin",
    "שיכוך כאב": "pain", "אנדוקרינולוגיה": "endocrine", "חיסונים": "vaccine", "תוספי תזונה": "nutrition",
    "נוירולוגיה": "neuro", "חטיפים": "nutrition", "לב ולחץ דם": "cardio", "פסיכיאטריה והרגעה": "neuro",
    "נשימה / שיעול": "respiratory", "הרדמה": "anesthesia", "סטרואידים": "pain",
}
KW_TOPIC = [
    ("ears", ("אוזני", "אוזן", "otic", "אקרי", "אזני", "auris", "ear ")),
    ("eyes", ("עיני", "עין ", "ophth", "ocular", "דמעות", "eye", "lash", "idrop", "tear")),
    ("dental", ("שיני", "דנטל", "dental", "אבנית", "פה ", "oral", "breath", "teeth", "orocare", "oratene")),
    ("joints", ("מפרק", "גלוקוזאמין", "כונדרו", "joint", "ניידות", "wejoint", "mobility", "cartil")),
    ("kidney", ("כליה", "כלייתי", "renal", "שתן", "urinary", "flutd")),
    ("skin", ("עור", "פרוה", "פרווה", "דרמט", "אלרגי", "derma", "skin", "שמפו", "shampoo", "topical", "paw", "keto-c", "zoo ")),
    ("cardio", ("~לב", "קרדי", "cardio", "cardiac")),
    ("endocrine", ("סוכרת", "אינסולין", "תריס", "הורמון", "diabet", "thyro")),
    ("parasites", ("פרעוש", "קרצי", "flea", "tick", "נקסגארד", "ברוולין", "פרונטליין", "סימפריקה", "ברווקטו")),
    ("deworm", ("תולעים", "תילוע", "דרונטל", "worm", "dewor", "מילבמקס")),
    ("vaccine", ("חיסון", "תרכיב", "vaccin")),
    ("antibiotic", ("אנטיביו", "amoxi", "אמוקסי", "ציפרו", "doxy", "cephal", "clav", "metronid", "מטרוניד")),
    ("anesthesia", ("הרדמה", "anesth", "sedat", "propofol", "פרופופול", "isoflur", "קטמין", "ketamin")),
    ("pain", ("meloxi", "מלוקסי", "carprof", "רימדיל", "gabapent", "nsaid", "כאב")),
    ("respiratory", ("שיעול", "נשימ", "cough", "bronch")),
    ("nutrition", ("מזון", "food", "פחית", "שק ", 'ק"ג', "ויטמין", "תוסף", "חטיף", "diet")),
    ("equipment", ("מזרק", "מחט", "כפפ", "חוט", "syring", "needle", "glove", "sutur", "קטטר", "cathet",
                   "תחבוש", "gauze", "גזה", "swab", "bandag", "catgut", "clinisorb", "clinisolv", "splint",
                   "microchip", "reader", "test", "pill", "cutanplast", "collar", "מתלה", "display")),
]
LAB_CAT_TOPIC = {
    "ביוכימיה": "chemistry", "פאנלים": "chemistry", "בדיקות כלליות": "chemistry", "פאנלים - חתולים": "chemistry",
    "פאנלים - כלבים": "chemistry", "פאנלים - עופות": "chemistry", "כליות": "chemistry", "לבלב": "chemistry",
    "לב": "chemistry", "GI": "chemistry", "דלקת": "chemistry",
    "המטולוגיה": "hematology", "ספירת דם": "hematology",
    "הורמונים": "endocrine", "תיירואיד": "endocrine",
    "סרולוגיה": "serology", "זיהומים": "serology", "חיסון": "serology", "סרולוגיה/PCR": "serology",
    "PCR": "pcr", "תרביות": "culture", "פתולוגיה": "pathology", "ציטולוגיה": "pathology",
    "שתן/צואה/תרביות": "urine", "שתן": "urine", "קרישה": "coag",
    "רמת תרופות בדם": "drugs", "תרופות": "drugs", "בדיקות מיוחדות": "other",
    "מחלות גנטיות": "pcr", "פתוגנים": "pcr", "פאנלים": "pcr",
    "Cytology": "pathology", "Histology": "pathology", "Immunohistochemistry": "pathology",
    "Ear Screens": "culture", "Allergy – Screens and Panels": "serology", "General": "chemistry",
    "Biochemistry": "chemistry", "Adrenal / Pituitary": "endocrine", "Thyroid": "endocrine",
    "Reproduction": "endocrine", "Other Hormones": "endocrine", "Infectious Disease": "serology",
    "Genetic Disease PCR": "pcr",
}

def _hit(kw, name, low):
    if kw.startswith("~"): return re.search(r"(?<![א-ת])" + kw[1:] + r"(?![א-ת])", name) is not None
    return kw in name or kw in low

def kw_topic(name):
    low = name.lower()
    for t, kws in KW_TOPIC:
        if any(_hit(k, name, low) for k in kws): return t
    return "other"

# Vet-diet food names carry the indication in the product line name (Renal / רינאל / Gastro …).
FOOD_KW = [
    ("kidney", ("רינאל", "renal", "c/d ", "k/d ", "s/d ", "u/d ", "multicare", "urinary", "אורינרי", "שתן", "כליה", "s/o", "uc ", "ct urin")),
    ("gi", ("גסטרו", "gastro", "i/d ", "gi biome", "w/d ", "intestinal", "אנטריק", "enteric", "digest", "פיברה", "fibre", "fiber",
            "ריקברי", "recovery", "קונבלסנס", "convalescence", "דל שומן", "low fat", "en ", "ha ", "hypo")),
    ("liver", ("הפאטיק", "hepatic", "l/d ", "כבד", "hp ")),
    ("skin", ("דרמו", "derma", "derm", "דרם", "d/d ", "z/d ", "סקין", "skin", "אלרג", "allerg", "sensitiv", "סנסיטיב", "אטופיק", "atopic", "hf ")),
    ("joints", ("מוביליטי", "mobility", "j/d ", "מפרק", "joint", "ja ")),
    ("dental", ("דנטל", "dental", "t/d ", "שיניים", "oral", "ds ")),
    ("endocrine", ("דיאבטיק", "diabetic", "m/d ", "y/d ", "סוכרת", "glycobalance", "dm ", "היפרתירואיד", "thyro")),
    ("cardio", ("קרדיאק", "cardiac", "~לב", "ck ")),
    ("neuro", ("neuro", "b/d ", "אפילפ", "epilep", "calm", "קאלם")),
    ("onco", ("onco", "on care", "onc care", "סרטן", "tumor")),
]
LAB_NAME_KW = [
    ("hematology", ("cbc", "hematolog", "blood smear", "reticuloc", "coombs", "blood type", "packed cell")),
    ("coag", ("coagul", "pt/", "aptt", "fibrinogen", "d-dimer", "clotting")),
    ("endocrine", ("thyro", "t4", "tsh", "cortisol", "acth", "insulin", "progester", "estrad", "testoster", "hormone", "ldds", "hdds")),
    ("serology", ("serolog", "antibod", " ab ", "titer", "titre", "elisa", "ifa", "leishman", "ehrlich", "anaplasm", "borrelia", "felv", "fiv", "heartworm", "dirofil", "toxoplasm", "brucell", "distemper", "parvo")),
    ("pcr", ("pcr", "realpcr", "genetic", "dna", "sequenc")),
    ("culture", ("culture", "sensitiv", "fungal", "dermatophyt", "aerobic", "anaerobic", "mic ")),
    ("pathology", ("cytolog", "histopath", "biopsy", "pathol", "aspirat", "fnа", "fna ", "necrops")),
    ("urine", ("urin", "fecal", "feces", "stool", "giardia", "upc", "sediment")),
    ("drugs", ("phenobarb", "level", "digoxin", "bromide", "cyclospor", "drug ")),
]
def LAB_NAME_TOPIC(name):
    low = name.lower()
    for t, kws in LAB_NAME_KW:
        if any(k in low for k in kws): return t
    return "chemistry"

def food_topic(name):
    low = name.lower()
    for t, kws in FOOD_KW:
        if any(_hit(k, name, low) for k in kws): return t
    return "nutrition"

# ---------------------------------------------------------------- food facets (food section only)
# A vet can filter the food section the way they think about it: therapeutic vs everyday, wet vs
# dry, life stage, species, and dog size. Size classes come from the brands' own naming (mini /
# small / toy vs medium / maxi / large), which is an approximation of Gil's 7 kg cut-off.
FOOD_FACETS = {
    "kind":    [("vet", "וטרינרי / רפואי"), ("regular", "רגיל")],
    "form":    [("dry", "יבש"), ("wet", "רטוב")],
    "stage":   [("puppy", "גורים"), ("adult", "בוגר"), ("senior", "מבוגר")],
    "dogsize": [("small", 'עד 7 ק"ג'), ("large", 'מעל 7 ק"ג')],
}
VET_FOOD = {"rc-vet", "hills-pd", "hills-ve", "purina-vet", "monge-vet", "vetlife"}
WET_KW = ("פחית", "פאוץ", "רטוב", "שימור", "מוס", "נזיד", "stew", "טרין", "ג'לי", "pouch",
          "loaf", "gravy", "רוטב", "מרק", "נוזל")
STAGE_KW = [("senior", ("מבוגר", "סניור", "senior", "mature", "aging", "אייג'ינג", "אייגינג", "7+", "8+", "11+", "12+")),
            ("puppy", ("גור", "פאפי", "puppy", "junior", "kitten", "קיטן", "starter", "סטרטר",
                       "אמהות", "mother", "growth", "ג'וניור", "גוניור")),
            ("adult", ("בוגר", "אדולט", "adult"))]
SIZE_KW = [("small", ("mini", "מיני", "small", "סמול", "toy", "טוי", "x-small", "xsmall",
                      "גזע קטן", "מגזע קטן", "קטן")),
           ("large", ("maxi", "מקסי", "large", "לארג", "giant", "ג'יינט", "medium", "מדיום",
                      "בינוני", "midi", "מידי", "גזע גדול", "גדול", "ג'יאנט", "גיאנט"))]

def _kw(hay, kws): return next((t for t, ks in kws if any(k in hay for k in ks)), None)

def food_facets(slug, name, unit, animal, category):
    hay = f"{name} {category or ''}".lower()
    wet = any(k in hay for k in WET_KW) or (unit or "").endswith("ג'")
    return {"kind": "vet" if slug in VET_FOOD else "regular",
            "form": "wet" if wet else "dry",
            "stage": _kw(hay, STAGE_KW),
            "dogsize": _kw(hay, SIZE_KW) if animal != "חתול" else None}

def food_animal(name, category):
    hay = f"{name} {category or ''}".lower()
    if any(k in hay for k in ("חתול", "קיטן", "cat", "feline", "kitten")): return "חתול"
    if any(k in hay for k in ("כלב", "פאפי", "dog", "canine", "puppy")): return "כלב"
    return None

# ---------------------------------------------------------------- price-list registry (dates + sources = manual truth)
DLD = HOME / "Downloads"
REG = {
    # slug: (type, supplier label, price_list_date, source_file_path or None, vat_basis of the raw list, notes)
    "beit-erez":  ("medical", "בית ארז (מילטין)", "2026-07", DLD / "מחירון - חיות קטנות יולי 2026.pdf", "no_vat", "מחירון חיות קטנות יולי 2026"),
    "msd":        ("medical", "MSD (ברווקטו)", "2026", HOME / "pricecmp" / "pdf" / "new_bravecto.pdf", "no_vat", "מחיר לווטרינר 2026"),
    "miltin-consum": ("medical", "בית ארז — ציוד מתכלה", "2025-11", DLD / "PDF" / "מחירון חטיבה וטרינרית קבוצת מילטין ציוד מתכלה - נובמבר 2025.pdf", "no_vat", "חטיבה וטרינרית מילטין — ציוד מתכלה נובמבר 2025"),
    "zoetis":     ("medical", "זואטיס (Zoetis)", "2026", (DLD / "זואטיס תרופות 2026.PDF", DLD / "זואטיס סימפריקה סטרונגהולד 2026.PDF"), "no_vat", "מחיר לווטרינר 2026 — תרופות + סימפריקה/סטרונגהולד"),
    "kong":       ("medical", "קונג (Kong)", "2026-08", DLD / "מחירון קונג מלאי - אוגוסט 2026 .pdf", "no_vat", "מחירון מלאי אוגוסט 2026 — צעצועים ואביזרים"),
    "ferplast":   ("medical", "פרפלסט (Ferplast)", "2026-02", DLD / "מחירון מוצרי פרפלסט 24.2.26.pdf", "no_vat", "מחירון 24.2.2026 — מחיר מחירון (ללא הנחות)"),
    "vetmarket":  ("medical", "וטמרקט", "2026-08", None, "no_vat", "לוטמרקט אין מחירון מפורסם. הרשימה נבנית מפרסור 123 אישורי הזמנה (09/2023–08/2026) — מחיר המחירון לפני הנחה, עם תאריך לכל שורה"),
    "medi-market": ("medical", "מדי-מרקט", "2026-06", None, "with_vat", "מחירי אתר medi-market.co.il (נאספו 30/06/2026)"),
    "petvet":     ("medical", "פט-וט ביומד", "2026-05", DLD / "PRICE LIST 2026.pdf", "no_vat", "מחירון 2026 — כל המותגים (DermatoVet, Zymox, WePharm, VetInnov, Uranotest ועוד)"),
    "rc-vet":     ("food", "Royal Canin VET", "2026-06", DLD / "RC VET Price list JUNE.pdf", "no_vat", "מחירון קמעונאי יוני 2026"),
    "rc-retail":  ("food", "Royal Canin חנויות", "2026-01", DLD / "RC SPT Price list Jan_2026.pdf", "no_vat", "מחירון קמעונאי ינואר 2026"),
    "hills-pd":   ("food", "Hill's Prescription Diet", "2026-04", DLD / "PD_priceList_Apr26.PDF", "no_vat", "מחירון אפריל 2026 — מק\"ט ומשקל אריזה מהמחירון עצמו"),
    "hills-sp":   ("food", "Hill's Science Plan", "2026-04", DLD / "SP_PriceList_Apr26_2.PDF", "no_vat", "מחירון אפריל 2026 (בתוקף ממאי 2026)"),
    "hills-ve":   ("food", "Hill's Vet Essentials", "2026-04", None, "no_vat", None),
    "vetlife":    ("food", "VetLife", "2025-02", DLD / "PDF" / "מחירון וטלייף 02.25 (1).pdf", "no_vat", "מחירון פברואר 2025"),
    "purina-vet": ("food", "Purina Pro Plan VET", "2026-06", DLD / "פורינה.pdf", "no_vat", "מחיר מחירון ליחידה (ללא הנחות)"),
    "purina-retail": ("food", "פורינה (חנויות)", "2026-06", DLD / "הצעת מחיר וטרינר 2026.pdf", "no_vat", "פריסקיז · פנסי פיסט · גורמה · פרו פלאן · דנטלייף — מחיר מחירון בלבד"),
    "monge-vet":  ("food", "Monge Vet Solution", "2026-07", DLD / "מחירון מונג וט סלושיין יבש+רטוב - יולי 2026.pdf", "no_vat", "מחירון יולי 2026 — יבש + רטוב"),
    "monge":      ("food", "Monge", "2026-07", (DLD / "מחירון מונג יבש - יולי 2026.pdf", DLD / "מחירון מונג ביווילד - יולי 2026.pdf"), "no_vat", "מחירון יולי 2026 — יבש + BeWild"),
    "foodiez":    ("food", "Foodiez", None, None, "no_vat", None),
    "aml":        ("labs", "AML", "2026-06", DLD / "מחירון 2026.pdf", "with_vat", "בתוקף מ-1.6.2026"),
    "hamachon":   ("labs", "המכון (מעבדה חיצונית)", "2026-01", DLD / "מחירון מעבדה חיצונית לשנת 2026 (1) (1).docx", "no_vat", None),
    "idexx":      ("labs", "IDEXX", "2026-01", DLD / "מחירון איידקס מקוצר 2026xlsx.pdf", "no_vat", "מחירון מקוצר 2026 — הבדיקות המוצעות בישראל"),
    "idexx-ref":  ("labs", "IDEXX — רפרנס", "2025-01", DLD / "PDF" / "מחירון רפרנס איידקס 2025.pdf", "no_vat", "מחירון רפרנס 2025 — בדיקות שאינן במחירון המקוצר 2026; חלקן אינן זמינות בישראל"),
    "karnieli":   ("labs", "קרניאלי", "2025-01", (DLD / "PDF" / "מחירון פאנלים 2025 (2).pdf",
                                                  DLD / "PDF" / "מחירון מחלות גנטיות 2025 (1).pdf",
                                                  DLD / "PDF" / "מחירון וטרינרים פתוגנים כלבים חתולים 2025 (5).pdf"), "no_vat",
                   "פאנלים · מחלות גנטיות · פתוגנים — שלושת מחירוני 2025"),
    "banet":      ("labs", "פרופ' בנעט", None, None, "no_vat", None),
}
FOOD_COMPANY_SLUG = {"RC VET": "rc-vet", "RC חנויות": "rc-retail", "Hill's PD": "hills-pd", "Hill's VE": "hills-ve",
                     "VetLife": "vetlife", "Purina": "purina-vet", "Purina חנויות": "purina-retail",
                     "Monge Vet": "monge-vet", "Monge": "monge", "Foodiez": "foodiez"}
FROM_PDF = {"rc-vet", "rc-retail", "purina-vet", "purina-retail", "vetlife", "monge", "monge-vet", "hills-pd"}   # dated PDF replaces the undated TS rows
LAB_SLUG = {"AML": "aml", "המכון": "hamachon", "IDEXX": "idexx", "קרניאלי": "karnieli", "פרופ בנעט": "banet"}

# ---------------------------------------------------------------- shop section (non-food, non-medical)
SHOP_TOPICS = [("treats", "חטיפים"), ("toys", "צעצועים"), ("grooming", "טיפוח והיגיינה"),
               ("litter", "חול ושירותים"), ("accessories", "אביזרים"), ("calming", "הרגעה"),
               ("other", "כללי")]
SHOP_KW = [
    ("toys", ("צעצוע", " toy", "~כדור", " ball", "~חבל", "rope", "forager", "boredom", "~משחק")),
    ("treats", ("~חטיף", "~חטיפים", "snack", "מקלון", "עצם ללעיסה", "dentalife", "דנטלייף", "party mix",
                "פארטי מיקס", "שלוקים", "dental care", "דנטלקייר", "dogli", "דוגלי", "treat")),
    ("litter", ("~חול", "litter", "שירותים לחתול", "מגרפ", "scoop", "ארגז חול", "מצע")),
    ("grooming", ("מגבונ", "wipes", "שמפו", "shampoo", "מברשת", "brush", "טיפוח", "groom", "מסרק",
                  "גוזם", "ציפורניים", "nail")),
    ("calming", ("pet remedy", "מפיץ ריח", "atomizer", "בנדנה", "bandana", "פרומון", "pheromon")),
    ("accessories", ("~קערה", " bowl", "רצועה", "leash", "מנשא", "carrier", "~מיטה", " bed ",
                     "דיספלי", "display")),
]
SHOP_SUPPLIERS = {"ferplast", "kong"}   # whole list is shop merchandise
SHOP_FALLBACK = {"ferplast": "accessories", "kong": "toys"}   # sub-category when no keyword matches                 # whole list is shop merchandise
SHOP_BRANDS = {"Pet Remedy"}                  # brand inside a mixed list
STRONG_SHOP = {"treats", "toys", "litter"}    # unambiguous merchandise — always moves
CLINICAL = {"ears", "eyes", "dental", "skin", "joints", "kidney", "gi", "cardio", "endocrine", "neuro",
            "liver", "onco", "antibiotic", "pain", "parasites", "deworm", "vaccine", "respiratory",
            "anesthesia", "equipment"}

def shop_cat(it, slug):
    """-> shop sub-category key, or None if the row stays in its own section.
    A clinical row (medicated shampoo, ear wipes, HMEF filter) only moves for unambiguous
    merchandise categories, so vet products don't leak into the shop section."""
    c = _shop_kw(it)
    if slug in SHOP_SUPPLIERS: return c or SHOP_FALLBACK[slug]
    if (it.get("category") or "") in SHOP_BRANDS: return c or "calming"
    if not c: return None
    if c in STRONG_SHOP: return c
    return c if it.get("topic") not in CLINICAL else None

def _shop_kw(it):
    hay = f"{it.get('name','')} {it.get('category') or ''}"
    low = hay.lower()
    for t, kws in SHOP_KW:
        if any(_hit(k, hay, low) for k in kws): return t
    return None

# What each list still needs — shown on the "מצב המחירונים" page so it is obvious what to chase.
ACTIONS = {
    "idexx-ref": ("refresh", "מחירי 2025. כל הבדיקות שחוזרות במחירון המקוצר 2026 הועברו לשם — כאן נשארו רק בדיקות שלא מופיעות בו."),
    "hills-ve": ("no_source", "אין מחירון PDF — 42 שורות מקובץ המרפאה. ייתכן שהקו הוחלף ב-Science Plan; כדאי לאמת מול הספק."),
    "banet":    ("no_source", "אין קובץ מחירון בידינו — 23 בדיקות מקובץ המרפאה בלבד."),
    "medi-market": ("ok", "נאסף אוטומטית מאתר medi-market.co.il."),
    "vetmarket": ("ok", "אין מחירון מפורסם לוטמרקט — כל שורה היא מחיר המחירון לפני הנחה מתוך אישור הזמנה, עם התאריך שלה."),
}
ACTION_DEFAULT = {"current": ("ok", "מעודכן."),
                  "stale": ("refresh", "המחירון ישן מ-12 חודשים — כדאי לבקש מהספק מחירון עדכני."),
                  "missing_source": ("no_source", "אין קובץ מחירון מקורי בידינו.")}

def _hkey(s):
    """name+pack-size key that survives quote/order noise, for matching Hill's TS rows to PDF rows."""
    s = s.lower().replace('"', "").replace("'", "").replace("&", "")
    return re.sub(r"[^a-z0-9א-ת./]", "", s)

def r2(x): return round(float(x) + 1e-9, 2)

def item(slug, name, no_vat, with_vat, category, topic, **kw):
    if no_vat is None and with_vat is not None: no_vat = with_vat / VAT
    if with_vat is None and no_vat is not None: with_vat = no_vat * VAT
    if not no_vat or no_vat < 0.1: return None  # 0 / 0.01 = placeholder, not a price
    with_vat = no_vat * VAT  # single 18% basis everywhere (raw lists mix 17%/18%)
    d = {"name": name.strip(), "category": (category or "").strip() or None, "topic": topic,
         "price_no_vat": r2(no_vat), "price_with_vat": r2(with_vat)}
    for k, v in kw.items():
        if v not in (None, "", []): d[k] = v
    return d

def add(lst, it):
    if it is not None: lst.append(it)

def load(name): return json.load(open(CAN / name, encoding="utf-8"))

def pack_qty(name):
    m = re.search(r"(\d+)\s*(טבל|קפס|כדור|אמפול|פיפט|יח|tab|cap|caps|amp|unit|pcs|ml|מ\"ל|ml)", name, re.I)
    return int(m.group(1)) if m else None

def build():
    lists = {s: [] for s in REG}
    # --- Beit Erez (list price, ex-VAT, bonuses are supplier-published quantity bonuses = public)
    for it in load("beit_erez.json"):
        cat = BE_CANON.get(it["category"], it["category"])
        topic = BE_TOPIC.get(cat) or ("equipment" if any(k in cat for k in ("מעבדה", "בדיקות", "חוטים", "כפפות", "ניתוח", "טובוס", "מזרק", "מחט", "נוזל", "עירוי", "פרפר", "קטטר", "קולר", "חביש")) else kw_topic(it["item_name"]))
        bonus = " / ".join(b for b in (it.get("bonus1"), it.get("bonus2"), it.get("bonus3")) if b)
        notes = re.sub(r"\s*·\s*בונוס:.*$", "", it.get("notes") or "")
        add(lists["beit-erez"], item("beit-erez", it["item_name"], it["price_no_vat"], None, cat, topic,
                                       sku=it.get("code"), animal=it.get("animal"), manufacturer=it.get("manufacturer"),
                                       bonus=bonus, notes=notes, pack_qty=pack_qty(it["item_name"])))
    # --- Vetmarket. There is NO published Vetmarket price list, so every price here comes from
    #     parsing the order confirmations: unit_price is the list price BEFORE discount, and each
    #     row keeps the date of the confirmation it came from. The clinic's own vetmarket catalog
    #     (vetmarketCatalog.ts / "מחירון וטמרקט מלא.xlsx") is an internal file of unverifiable
    #     provenance — it is used ONLY to label a category, never for a price.
    pull = load("server_pull.json")
    vm_cat = {it.get("code"): it for it in load("vetmarket.json")}
    for r in pull["vetmarket_invoice_list"]:
        ref = vm_cat.get(r["sku"]) or {}
        cat = r.get("category") or ref.get("category")
        add(lists["vetmarket"], item("vetmarket", r["name"], r["unit_price"], None, cat,
                                       VM_TOPIC.get(cat) or kw_topic(r["name"]), sku=r["sku"],
                                       pack_qty=pack_qty(r["name"]), price_date=r["date"][:7]))
    # --- Pet-Vet Biomed: full 2026 price list (PDF, all brands). category = brand.
    for it in load("importer_2026.json"):
        add(lists["petvet"], item("petvet", it["name"], it["price_no_vat"], None, it.get("category"),
                                    kw_topic(it["name"]), sku=it.get("sku"), bonus=it.get("bonus"),
                                    pack_qty=pack_qty(it["name"])))
    # --- Zoetis 2026 (two sheets: meds + parasiticides)
    for it in load("zoetis_2026.json"):
        cat = it.get("category") or ""
        add(lists["zoetis"], item("zoetis", it["name"], it["price_no_vat"], None, cat,
                                    kw_topic(cat) if kw_topic(cat) != "other" else kw_topic(it["name"]),
                                    pack_qty=pack_qty(it["name"])))
    # --- MSD Bravecto 2026
    for it in load("msd_2026.json"):
        add(lists["msd"], item("msd", it["name"], it["price_no_vat"], None, it.get("category"), "parasites",
                                 sku=it.get("sku")))
    # --- Beit Erez / Miltin consumables (Nov 2025)
    for it in load("miltin_consum_2025_11.json"):
        cat = it.get("category") or ""
        add(lists["miltin-consum"], item("miltin-consum", it["name"], it["price_no_vat"], None, cat,
                                           BE_TOPIC.get(cat) or (kw_topic(it["name"]) if kw_topic(it["name"]) != "other" else "equipment"),
                                           sku=it.get("sku"), manufacturer=it.get("manufacturer"), notes=it.get("notes")))
    # --- Ferplast (equipment/accessories); the "25% off" column in the source is dropped on purpose.
    for it in load("ferplast_2026_02.json"):
        add(lists["ferplast"], item("ferplast", it["name"], it["price_no_vat"], None, it.get("category"),
                                      kw_topic(it["name"]) if kw_topic(it["name"]) != "other" else "equipment",
                                      sku=it.get("sku")))
    # --- Kong (toys/merchandise) — whole list lands in the shop section.
    for it in load("kong_2026_08.json"):
        add(lists["kong"], item("kong", it["name"], it["price_no_vat"], None, "Kong",
                                  kw_topic(it["name"]), sku=it.get("sku")))
    # --- Medi-Market: PUBLIC WEBSITE prices (regular_price, incl. VAT). medimarket_ts.json is NOT used (negotiated prices).
    for r in pull["medimarket"]:
        cl = json.loads(r["categories"] or "[]"); name = r["name"].replace("&quot;", '"').replace("&amp;", "&")
        add(lists["medi-market"], item("medi-market", name, None, r["regular_price"], cl[0] if cl else None,
                                         kw_topic(name), sku=r["sku"], pack_qty=pack_qty(name)))
    # --- Food from the clinic TS catalog: purchasePriceNoVat only (customerPriceWithVat = clinic sale
    #     price, deliberately dropped). Companies that now have a dated PDF are skipped here.
    for it in load("vet_food_catalog.json"):
        slug = FOOD_COMPANY_SLUG[it["company"]]
        if slug in FROM_PDF: continue
        add(lists[slug], item(slug, it["name"], it["purchasePriceNoVat"], None, it["indication"],
                                FOOD_IND_TOPIC.get(it["indication"], "nutrition"), animal=it.get("animal"), unit=it.get("weight")))
    # --- Food from dated supplier PDFs (list price, ex-VAT) — replaces the undated TS rows above.
    # Hill's PD/SP come straight from the April-2026 price-list PDF (SKU + pack weight included).
    # The clinic TS rows were ~3.5% lower across the board, so the PDF replaces them; only the
    # food indication (Renal / Gastro / …) is carried over from the TS by name+size.
    hills_ind = {_hkey(i["name"] + " " + (i.get("weight") or "")): i["indication"]
                 for i in load("vet_food_catalog.json") if i["company"] in ("Hill's PD", "Hill's VE")}
    for slug, src in (("hills-pd", "hills_pd_2026_04.json"), ("hills-sp", "hills_sp_2026_04.json")):
        for it in load(src):
            ind = hills_ind.get(_hkey(it["name"]))
            add(lists[slug], item(slug, it["name"], it["price_no_vat"], None, ind,
                                    FOOD_IND_TOPIC.get(ind) or food_topic(it["name"]),
                                    sku=it.get("sku"), unit=it.get("unit"),
                                    animal="חתול" if "חתול" in it["name"] else ("כלב" if "כלב" in it["name"] else None)))
    for slug, src in (("rc-vet", "rc_vet_2026_06.json"), ("rc-retail", "rc_spt_2026_01.json"),
                      ("purina-vet", "purina_2026_06.json"), ("purina-retail", "purina_retail_2026_06.json"),
                      ("vetlife", "vetlife_2025_02.json"), ("monge", "monge_2026_07.json"),
                      ("monge-vet", "monge_vet_2026_07.json")):
        for it in load(src):
            name = it["name"]
            add(lists[slug], item(slug, name, it["price_no_vat"], None, it.get("category"), food_topic(name),
                                    sku=it.get("sku"), unit=it.get("unit"),
                                    animal=it.get("animal") or ("חתול" if "חתול" in name else ("כלב" if "כלב" in name else None))))
    # --- IDEXX: the 2026 short list is the current one; the 2025 reference keeps only the tests
    #     it does not cover (118 of the 119 shared codes had moved 5-14%, so no 2025 price stands).
    def lab_topic(cat, name): return LAB_CAT_TOPIC.get((cat or "").replace("\n", " ")) or LAB_NAME_TOPIC(name)
    idexx_2026 = load("idexx_2026.json")
    for it in idexx_2026:
        add(lists["idexx"], item("idexx", it["name"], it["price_no_vat"], None, it.get("category"),
                                   lab_topic(it.get("category"), it["name"]), sku=it.get("sku"),
                                   notes=it.get("notes") or it.get("sample")))
    refreshed = {it["sku"] for it in idexx_2026}
    for it in load("idexx_2025.json"):
        if it.get("sku") in refreshed: continue
        add(lists["idexx-ref"], item("idexx-ref", it["name"], it["price_no_vat"], None, None,
                                       LAB_NAME_TOPIC(it["name"]), sku=it.get("sku"), notes=it.get("notes")))
    # --- Karnieli: panels + genetic diseases + pathogens, from the three 2025 PDFs
    KARN_SRC = {"פאנלים": "sources/labs/karnieli-2025-01.pdf",
                "מחלות גנטיות": "sources/labs/karnieli-2025-01-2.pdf",
                "פתוגנים": "sources/labs/karnieli-2025-01-3.pdf"}
    for it in load("karnieli_2025.json"):
        add(lists["karnieli"], item("karnieli", it["name"], it["price_no_vat"], None, it.get("category"),
                                      lab_topic(it.get("category"), it["name"]), notes=it.get("notes"),
                                      source=KARN_SRC.get(it.get("category"))))
    # --- Labs
    for it in load("lab_catalog.json"):
        slug = LAB_SLUG[it["lab"]]
        if slug in ("idexx", "karnieli"): continue      # both now come from their own price lists
        add(lists[slug], item(slug, it["name"], it.get("price_no_vat"), it.get("price_with_vat"), it["category"],
                                LAB_CAT_TOPIC.get(it["category"], "other"), sku=it.get("code"), notes=it.get("notes")))

    # --- food facets: therapeutic vs everyday, wet/dry, life stage, species, dog size
    for slug in [k for k in lists if REG[k][0] == "food"]:
        for it in lists[slug]:
            if not it.get("animal"):
                a = food_animal(it["name"], it.get("category"))
                if a: it["animal"] = a
            for k, v in food_facets(slug, it["name"], it.get("unit"), it.get("animal"), it.get("category")).items():
                if v: it[k] = v

    # --- move shop merchandise (non-food, non-medical) into its own section
    shop, SHOP_LABEL = {}, dict(SHOP_TOPICS)
    for slug in list(lists):
        if REG[slug][0] == "labs": continue
        keep = []
        for it in lists[slug]:
            c = shop_cat(it, slug)
            if c:
                it = dict(it, category=SHOP_LABEL[c], topic=c)
                shop.setdefault(slug, []).append(it)
            else:
                keep.append(it)
        lists[slug] = keep
    for slug, items in shop.items():
        if len(items) < 3:                       # 1-2 strays aren't a shop list — leave them in place
            lists[slug] += items; continue
        typ, label, date, src, vat, note = REG[slug]
        REG[f"{slug}-shop"] = ("shop", label, date, src, vat, note)
        lists[f"{slug}-shop"] = items

    # --- write
    for d in ("food", "medical", "labs", "shop"):
        (DATA / d).mkdir(parents=True, exist_ok=True)
        for old in (DATA / d).glob("*.json"): old.unlink()      # drop stale lists from earlier runs
    index = []
    for slug, (typ, label, date, src, vat_basis, note) in REG.items():
        items = lists[slug]
        if not items: continue                           # nothing priced -> not a price list
        for i, it in enumerate(items, 1): it["id"] = f"{slug}-{i}"
        src_rel, src_all = None, []
        if src:
            paths = list(src) if isinstance(src, tuple) else [src]
            rels = []
            for i, sp in enumerate(paths):
                if not sp.exists(): continue
                (SRC / typ).mkdir(parents=True, exist_ok=True)
                dst = SRC / typ / f"{slug}-{date}{'' if i == 0 else f'-{i+1}'}{sp.suffix.lower()}"
                if not dst.exists(): shutil.copy2(sp, dst)
                rels.append(f"sources/{typ}/{dst.name}")
            src_rel = rels[0] if rels else None
            src_all = rels
        status = "missing_source" if not date else ("stale" if date < stale_cutoff() else "current")
        act, act_note = ACTIONS.get(baseslug(slug), ACTION_DEFAULT[status])
        meta = {"action": act, "action_note": act_note,
                "slug": slug, "type": typ, "supplier": label, "price_list_date": date, "source_file": src_rel,
                "source_files": src_all, "vat_basis": vat_basis, "vat_rate": 18, "item_count": len(items), "status": status,
                "imported_at": TODAY, "notes": note}
        json.dump({"meta": meta, "items": items}, open(DATA / typ / f"{slug}.json", "w", encoding="utf-8"), ensure_ascii=False, indent=0)
        index.append(meta)
    tax = {"topics": TOPICS, "lab_topics": LAB_TOPICS, "shop_topics": SHOP_TOPICS, "food_facets": FOOD_FACETS, "food_indications": sorted({i["category"] for s in FOOD_COMPANY_SLUG.values() for i in lists[s] if i["category"]})}
    json.dump({"built_at": TODAY, "vat_rate": 18, "pricelists": index, "taxonomy": tax}, open(DATA / "index.json", "w", encoding="utf-8"), ensure_ascii=False, indent=1)
    return lists, index

def baseslug(s): return s[:-5] if s.endswith("-shop") else s

def stale_cutoff():
    d = datetime.date.today() - datetime.timedelta(days=365)  # annual lists are normal; >12 months = stale
    return d.strftime("%Y-%m")

def xlsx(lists, index):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        print("openpyxl missing — skipping xlsx"); return
    out = ROOT / "downloads"; out.mkdir(parents=True, exist_ok=True)
    by_type = {}
    for m in index: by_type.setdefault(m["type"], []).append(m["slug"])
    heb = {"food": "מזון", "medical": "מוצרים ותרופות", "labs": "מעבדות", "shop": "מוצרי חנויות"}
    for typ, slugs in by_type.items():
        wb = Workbook(); ws = wb.active; ws.title = "הגדרות"; ws.sheet_view.rightToLeft = True
        ws.append(["ספק", "ההנחה שלי %", "מרווח %", "מרווח ₪ קבוע"])
        for c in ws[1]: c.font = Font(bold=True)
        sup_rows = {}
        for s in slugs:
            m = next(x for x in index if x["slug"] == s)
            ws.append([m["supplier"], 0, 0, 0]); sup_rows[s] = ws.max_row
        ws.append([]); ws.append(["נוסחה: עלות = מחיר ללא מע\"מ × (1 − הנחה) × 1.18 ; מחיר ללקוח = עלות × (1 + מרווח %) + מרווח ₪"])
        ws2 = wb.create_sheet(heb[typ]); ws2.sheet_view.rightToLeft = True
        FCOLS = [("kind", "סוג המזון"), ("form", "יבש / רטוב"), ("stage", "שלב חיים"),
                 ("animal", "חיה"), ("dogsize", "גודל הכלב")] if typ == "food" else []
        FHEB = {k: dict(v) for k, v in FOOD_FACETS.items()}
        hdr = ["ספק", "פריט", "קטגוריה", "מק\"ט"] + [h for _, h in FCOLS] + ["תאריך מחירון", "מחיר ללא מע\"מ", "מחיר כולל מע\"מ", "הנחה %", "עלות אחרי הנחה (כולל מע\"מ)", "מרווח %", "מרווח ₪", "מחיר ללקוח"]
        C = chr(ord("E") + len(FCOLS))          # first price column shifts by the facet columns
        cols = [chr(ord(C) + i) for i in range(8)]
        ws2.append(hdr)
        for c in ws2[1]: c.font = Font(bold=True); c.fill = PatternFill("solid", fgColor="DCEFEB")
        for s in slugs:
            m = next(x for x in index if x["slug"] == s); r = sup_rows[s]
            for it in lists[s]:
                row = ws2.max_row + 1
                p, d, c2, mp, mf, cust = cols[1], cols[3], cols[4], cols[5], cols[6], cols[7]
                ws2.append([m["supplier"], it["name"], it.get("category"), it.get("sku")]
                           + [FHEB.get(k, {}).get(it.get(k), it.get(k) or "") for k, _ in FCOLS]
                           + [m["price_list_date"], it["price_no_vat"], it["price_with_vat"],
                              f"=הגדרות!B{r}", f"={p}{row}*(1-{d}{row}/100)*1.18", f"=הגדרות!C{r}",
                              f"=הגדרות!D{r}", f"={c2}{row}*(1+{mp}{row}/100)+{mf}{row}"])
        widths = [18, 55, 20, 12] + [13] * len(FCOLS) + [12, 14, 14, 9, 18, 9, 9, 14]
        for i, w in enumerate(widths): ws2.column_dimensions[chr(ord("A") + i)].width = w
        wb.save(out / f"{typ}.xlsx")

if __name__ == "__main__":
    lists, index = build()
    xlsx(lists, index)
    for m in index: print(f'{m["type"]:8} {m["slug"]:14} {m["item_count"]:5} {m["price_list_date"] or "—":8} {m["status"]}')
    print("total", sum(m["item_count"] for m in index))
